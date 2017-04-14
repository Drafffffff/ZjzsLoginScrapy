# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook


path = os.getcwd() + '/Data/test.xlsx'
wb = load_workbook(filename=path)

ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])  # 获取特定的 worksheet


def get_name(key):
    tmp = 'A' + str(key)
    return ws[tmp].value


def get_identid(key):
    tmp = 'B' + str(key)
    return ws[tmp].value


# for i in xrange(1, 5):
#     print get_identid(i)

def get_passwd(key):
    tmp = 'C' + str(key)
    return ws[tmp].value

# for i in xrange(1, 5):
#     print get_passwd(i)


def wr_in(col, key, text):
    tmp = col + str(key)
    ws[tmp].value = text
    wb.save(path)


# ws['A5'].value = 100
# wb.save(path)


# for i in xrange(1,100):
#     wr_in('E', i, i)
